#!/usr/bin/env python
from __future__ import annotations

import argparse
import fnmatch
import json
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .model_registry import DEFAULT_MODEL_MATRIX, model_slug, parse_model_list
    from .reference_eval import expand_dataset_patterns, get_reference_datasets, raw_data_root
except ImportError:  # pragma: no cover - script execution path
    from model_registry import DEFAULT_MODEL_MATRIX, model_slug, parse_model_list
    from reference_eval import expand_dataset_patterns, get_reference_datasets, raw_data_root


HEADERS = [
    "Dataset",
    "Train/Test",
    "# of CGM files in Gold Standard",
    "# of CGM files in Reference Repository",
    "# of files found by INSIGHT",
    "# of files found by INSIGHT in Gold Standard",
    "# of files found by INSIGHT in Reference Repository",
    "# of non-Reference files found by INSIGHT",
    "# of Files that file Loader runs on for Reference Repository files",
    "# of Files that file Loader runs on for non-Reference Repository files",
    "# of files Subject ID correctly extracted for Reference Repository",
    "Reference Repository files extracted correctly",
    "# of files processed",
    "Dataset Subject Precision",
    "Dataset Subject Recall",
    "Dataset Subject F1",
    "Dataset Temporal Precision",
    "Dataset Temporal Recall",
    "Dataset Temporal F1",
    "Dataset Glucose MAE",
    "Dataset Glucose RMSE",
    "Dataset Within 5 mg/dL",
    "Dataset Within 10 mg/dL",
    "Dataset Within 20 mg/dL",
    "Dataset INSIGHT Score",
    "Dataset Robust Score",
    "Scalar Feature Score",
    "Scalar Mean Glucose Delta",
    "Scalar Time In Range Delta",
    "Scalar GMI Delta",
    "Dataset Assessment",
    "Benchmark Status",
    "Benchmark Reason",
    "Reference Coverage Status",
    "Reference Parse Status",
    "Reference Comparison Scope",
    "Notes",
    "Gold Standard Files",
    "Reference Repository Files",
    "INSIGHT Files",
    "INSIGHT Files extracted correctly from Reference Repository",
    "INSIGHT Files processed",
    "Gold TP",
    "Gold FP",
    "Gold FN",
    "Gold Precision",
    "Gold Recall",
    "Gold F1",
    "Reference TP",
    "Reference FP",
    "Reference FN",
    "Reference Precision",
    "Reference Recall",
    "Reference F1",
]

SUMMARY_HEADERS = [
    "Model",
    "Train/Test",
    "# of Active Datasets",
    "# of CGM Files in Gold Standard",
    "# of CGM Files in Reference Repository",
    "Gold File Precision",
    "Gold File Recall",
    "Gold File F1",
    "Reference File Precision",
    "Reference File Recall",
    "Reference File F1",
    "% of Datasets Extracted Correctly",
    "Avg Dataset Subject F1",
    "Avg Dataset Temporal F1",
    "Avg Dataset Within 10 mg/dL",
    "Avg Dataset INSIGHT Score",
    "Avg Dataset Robust Score",
    "Avg Scalar Feature Score",
    "Avg Dataset Glucose MAE",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="E7F4E4")


def _safe_rate(numerator: float, denominator: float) -> float:
    return float(numerator / denominator) if denominator else 0.0


def _safe_number(value) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def _mean(values: list[float]) -> float:
    return float(sum(values) / len(values)) if values else 0.0


def _f1(precision: Optional[float], recall: Optional[float]) -> Optional[float]:
    if precision is None or recall is None or not precision or not recall:
        return 0.0 if precision == 0 or recall == 0 else None
    return float(2 * precision * recall / (precision + recall))


def _nested_get(data: dict, *path: str):
    current = data
    for key in path:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def sheet_title_for_model(model_name: str) -> str:
    title = model_name
    if title.lower().startswith("gpt-"):
        title = "GPT-" + title[4:]
    return title


def _load_manifest(manifest_path: Path) -> list[dict]:
    if not manifest_path.exists():
        return []
    data = json.loads(manifest_path.read_text())
    return list(data.get("files", []))


def _rel_source(dataset_root: Path, source_path: str) -> str:
    source = Path(source_path)
    try:
        return source.relative_to(dataset_root).as_posix()
    except Exception:
        return source.name


def _matches_any(path_str: str, patterns: Iterable[str]) -> bool:
    rel = path_str.lower()
    return any(fnmatch.fnmatch(rel, pattern.lower()) for pattern in patterns)


def _summarize_paths(paths: list[str], *, label: str, max_items: int = 8) -> Optional[str]:
    if not paths:
        return None
    if len(paths) > max_items:
        return label
    basenames = [Path(path).name for path in paths]
    if len(set(basenames)) == len(basenames):
        return ", ".join(basenames)
    return ", ".join(paths)


def _load_comparison_metrics(comparison_json: Path) -> dict:
    if not comparison_json.exists():
        return {}
    return json.loads(comparison_json.read_text())


def _load_benchmark_status(benchmark_json: Path) -> dict:
    if not benchmark_json.exists():
        return {}
    return json.loads(benchmark_json.read_text())


def _comparison_dataset_metrics(comparison: dict) -> dict[str, object]:
    subject_precision = _nested_get(comparison, "subject_metrics", "precision")
    subject_recall = _nested_get(comparison, "subject_metrics", "recall")
    subject_f1 = _nested_get(comparison, "subject_metrics", "f1")
    if subject_f1 is None:
        subject_f1 = _f1(subject_precision, subject_recall)

    temporal_precision = _nested_get(comparison, "temporal_alignment_metrics", "precision")
    temporal_recall = _nested_get(comparison, "temporal_alignment_metrics", "recall")
    temporal_f1 = _nested_get(comparison, "temporal_alignment_metrics", "f1")
    if temporal_precision is None:
        temporal_precision = _nested_get(comparison, "row_metrics", "precision")
    if temporal_recall is None:
        temporal_recall = _nested_get(comparison, "row_metrics", "recall")
    if temporal_f1 is None:
        temporal_f1 = _nested_get(comparison, "row_metrics", "f1")

    glucose_mae = _nested_get(comparison, "glucose_agreement_metrics", "glucose_mae")
    if glucose_mae is None:
        glucose_mae = _nested_get(comparison, "aligned_subject_time_metrics", "glucose_mae")
    glucose_rmse = _nested_get(comparison, "glucose_agreement_metrics", "glucose_rmse")
    within_5 = _nested_get(comparison, "glucose_agreement_metrics", "within_5mgdl_rate")
    within_10 = _nested_get(comparison, "glucose_agreement_metrics", "within_10mgdl_rate")
    within_20 = _nested_get(comparison, "glucose_agreement_metrics", "within_20mgdl_rate")
    insight_score = _nested_get(comparison, "insight_metrics", "overall_score")
    robust_score = _nested_get(comparison, "robust_insight_metrics", "overall_score")
    assessment = _nested_get(comparison, "insight_metrics", "assessment")
    scalar_score = _nested_get(comparison, "scalar_cgm_feature_metrics", "scalar_feature_score")
    scalar_features = _nested_get(comparison, "scalar_cgm_feature_metrics", "feature_comparisons") or {}
    mean_glucose_delta = scalar_features.get("mean_glucose", {}).get("absolute_delta")
    time_in_range_delta = scalar_features.get("time_in_70_180_rate", {}).get("absolute_delta")
    gmi_delta = scalar_features.get("gmi", {}).get("absolute_delta")
    reference_scope = _nested_get(comparison, "reference_scope", "comparison_scope")
    reference_note = _nested_get(comparison, "reference_scope", "note")

    if assessment is None and comparison:
        assessment = "legacy_only"

    return {
        "subject_precision": subject_precision,
        "subject_recall": subject_recall,
        "subject_f1": subject_f1,
        "temporal_precision": temporal_precision,
        "temporal_recall": temporal_recall,
        "temporal_f1": temporal_f1,
        "glucose_mae": glucose_mae,
        "glucose_rmse": glucose_rmse,
        "within_5": within_5,
        "within_10": within_10,
        "within_20": within_20,
        "insight_score": insight_score,
        "robust_score": robust_score,
        "scalar_score": scalar_score,
        "mean_glucose_delta": mean_glucose_delta,
        "time_in_range_delta": time_in_range_delta,
        "gmi_delta": gmi_delta,
        "assessment": assessment,
        "reference_scope": reference_scope,
        "reference_note": reference_note,
    }


def _is_reference_extract_correct(comparison: dict) -> bool:
    if not comparison:
        return False

    subject_metrics = comparison.get("subject_metrics", {})
    temporal_metrics = comparison.get("temporal_alignment_metrics", {})
    glucose_metrics = comparison.get("glucose_agreement_metrics", {})
    if temporal_metrics or glucose_metrics:
        glucose_quality = glucose_metrics.get("within_10mgdl_rate")
        if glucose_quality is None:
            glucose_quality = glucose_metrics.get("exact_glucose_rate")
        return (
            subject_metrics.get("precision") == 1
            and subject_metrics.get("recall") == 1
            and (temporal_metrics.get("f1") or 0.0) >= 0.999
            and (glucose_quality or 0.0) >= 0.999
        )

    row_metrics = comparison.get("row_metrics", {})
    row_f1 = row_metrics.get("f1")
    return (
        row_f1 is not None
        and row_f1 >= 0.999
        and subject_metrics.get("precision") == 1
        and subject_metrics.get("recall") == 1
    )


def _build_notes(manifest_entries: list[dict], comparison: dict, benchmark_status: dict | None = None) -> Optional[str]:
    warnings = []
    for entry in manifest_entries:
        warnings.extend(entry.get("warnings") or [])

    parts: list[str] = []
    if warnings:
        parts.append(f"warnings={sorted(set(warnings))}")

    if comparison:
        metrics = _comparison_dataset_metrics(comparison)
        if metrics["insight_score"] is not None and metrics["insight_score"] < 0.999:
            parts.append(f"insight_score={metrics['insight_score']:.4f}")
        if metrics["temporal_f1"] is not None and metrics["temporal_f1"] < 0.999:
            parts.append(f"temporal_f1={metrics['temporal_f1']:.4f}")
        if metrics["subject_recall"] is not None and metrics["subject_recall"] < 1:
            parts.append(f"subject_recall={metrics['subject_recall']:.4f}")
        if metrics["within_10"] is not None and metrics["within_10"] < 0.999:
            parts.append(f"within10={metrics['within_10']:.4f}")
        if metrics["glucose_mae"]:
            parts.append(f"glucose_mae={metrics['glucose_mae']:.4f}")
        if metrics["scalar_score"] is not None and metrics["scalar_score"] < 0.999:
            parts.append(f"scalar_feature_score={metrics['scalar_score']:.4f}")
        if metrics["reference_note"]:
            parts.append(f"reference_scope_note={metrics['reference_note']}")

    if benchmark_status and benchmark_status.get("status") == "quarantined":
        parts.append("quarantined")
        if benchmark_status.get("reason"):
            parts.append(f"benchmark_reason={benchmark_status['reason']}")

    return "; ".join(parts) if parts else None


def _populate_model_sheet(
    ws,
    model_name: str,
    runs_root: Path,
    evaluation_root: Path,
    splits: list[str],
) -> dict[str, tuple[int, int]]:
    sheet_ranges: dict[str, tuple[int, int]] = {}
    header_col = {header: index for index, header in enumerate(HEADERS, start=1)}

    for header, col in header_col.items():
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    current_row = 2
    slug = model_slug(model_name)

    for split in ["testing", "training"]:
        if split not in splits:
            continue

        split_rows_start = current_row
        datasets = get_reference_datasets(split)
        for dataset_ref in datasets:
            dataset_root = raw_data_root(split) / dataset_ref.dataset
            gold_files = expand_dataset_patterns(dataset_ref, dataset_ref.gold_patterns)
            reference_files = expand_dataset_patterns(dataset_ref, dataset_ref.reference_patterns)

            manifest_path = runs_root / slug / split / dataset_ref.dataset / "manifest.json"
            manifest_entries = _load_manifest(manifest_path)

            source_files = [_rel_source(dataset_root, entry["source"]) for entry in manifest_entries]
            processed_entries = [entry for entry in manifest_entries if entry.get("clean")]
            processed_sources = [_rel_source(dataset_root, entry["source"]) for entry in processed_entries]

            found_in_gold = [path for path in source_files if _matches_any(path, dataset_ref.gold_patterns)]
            found_in_reference = [path for path in source_files if _matches_any(path, dataset_ref.reference_patterns)]
            non_reference_found = [path for path in source_files if not _matches_any(path, dataset_ref.reference_patterns)]
            reference_entries = [
                entry for entry in manifest_entries
                if _matches_any(_rel_source(dataset_root, entry["source"]), dataset_ref.reference_patterns)
            ]
            ref_subject_ok = [
                entry for entry in reference_entries
                if entry.get("subject_id_strategy") not in (None, "unknown")
            ]

            comparison_json = (
                evaluation_root
                / slug
                / "comparisons"
                / split
                / dataset_ref.dataset
                / "comparison.json"
            )
            benchmark_json = (
                evaluation_root
                / slug
                / "comparisons"
                / split
                / dataset_ref.dataset
                / "benchmark_status.json"
            )
            comparison = _load_comparison_metrics(comparison_json)
            benchmark_status = _load_benchmark_status(benchmark_json)
            comparison_metrics = _comparison_dataset_metrics(comparison)

            exact_reference_extract = (
                len(reference_files)
                if comparison and _is_reference_extract_correct(comparison)
                else 0
            )

            gold_tp = len(found_in_gold)
            gold_fp = max(len(source_files) - gold_tp, 0)
            gold_fn = max(len(gold_files) - gold_tp, 0)
            reference_tp = len(found_in_reference)
            reference_fp = max(len(source_files) - reference_tp, 0)
            reference_fn = max(len(reference_files) - reference_tp, 0)

            def set_value(header: str, value) -> None:
                ws.cell(current_row, header_col[header], value)

            def col_letter(header: str) -> str:
                return get_column_letter(header_col[header])

            set_value("Dataset", dataset_ref.display_name)
            set_value("Train/Test", "Test" if split == "testing" else "Train")
            set_value("# of CGM files in Gold Standard", len(gold_files))
            set_value("# of CGM files in Reference Repository", len(reference_files))
            set_value("# of files found by INSIGHT", len(source_files))
            set_value("# of files found by INSIGHT in Gold Standard", gold_tp)
            set_value("# of files found by INSIGHT in Reference Repository", reference_tp)
            set_value("# of non-Reference files found by INSIGHT", len(non_reference_found))
            set_value("# of Files that file Loader runs on for Reference Repository files", len(reference_entries))
            set_value("# of Files that file Loader runs on for non-Reference Repository files", len(non_reference_found))
            set_value("# of files Subject ID correctly extracted for Reference Repository", len(ref_subject_ok))
            set_value("Reference Repository files extracted correctly", exact_reference_extract)
            set_value("# of files processed", len(processed_entries))

            set_value("Dataset Subject Precision", comparison_metrics["subject_precision"])
            set_value("Dataset Subject Recall", comparison_metrics["subject_recall"])
            set_value("Dataset Subject F1", comparison_metrics["subject_f1"])
            set_value("Dataset Temporal Precision", comparison_metrics["temporal_precision"])
            set_value("Dataset Temporal Recall", comparison_metrics["temporal_recall"])
            set_value("Dataset Temporal F1", comparison_metrics["temporal_f1"])
            set_value("Dataset Glucose MAE", comparison_metrics["glucose_mae"])
            set_value("Dataset Glucose RMSE", comparison_metrics["glucose_rmse"])
            set_value("Dataset Within 5 mg/dL", comparison_metrics["within_5"])
            set_value("Dataset Within 10 mg/dL", comparison_metrics["within_10"])
            set_value("Dataset Within 20 mg/dL", comparison_metrics["within_20"])
            set_value("Dataset INSIGHT Score", comparison_metrics["insight_score"])
            set_value("Dataset Robust Score", comparison_metrics["robust_score"])
            set_value("Scalar Feature Score", comparison_metrics["scalar_score"])
            set_value("Scalar Mean Glucose Delta", comparison_metrics["mean_glucose_delta"])
            set_value("Scalar Time In Range Delta", comparison_metrics["time_in_range_delta"])
            set_value("Scalar GMI Delta", comparison_metrics["gmi_delta"])
            set_value("Dataset Assessment", comparison_metrics["assessment"])
            set_value("Benchmark Status", benchmark_status.get("status", "active"))
            set_value("Benchmark Reason", benchmark_status.get("reason"))
            set_value("Reference Coverage Status", benchmark_status.get("reference_coverage_status"))
            set_value("Reference Parse Status", benchmark_status.get("reference_parse_status"))
            set_value("Reference Comparison Scope", comparison_metrics["reference_scope"])

            set_value("Notes", _build_notes(manifest_entries, comparison, benchmark_status))
            set_value("Gold Standard Files", dataset_ref.gold_label)
            set_value("Reference Repository Files", dataset_ref.reference_label)
            set_value(
                "INSIGHT Files",
                _summarize_paths(
                    source_files,
                    label=dataset_ref.gold_label if len(found_in_gold) == len(source_files) else dataset_ref.reference_label,
                ),
            )
            set_value(
                "INSIGHT Files extracted correctly from Reference Repository",
                dataset_ref.reference_label if exact_reference_extract == len(reference_files) and reference_files else None,
            )
            set_value("INSIGHT Files processed", _summarize_paths(processed_sources, label=dataset_ref.gold_label))

            set_value("Gold TP", gold_tp)
            set_value("Gold FP", gold_fp)
            set_value("Gold FN", gold_fn)
            set_value("Reference TP", reference_tp)
            set_value("Reference FP", reference_fp)
            set_value("Reference FN", reference_fn)

            gold_tp_col = col_letter("Gold TP")
            gold_fp_col = col_letter("Gold FP")
            gold_fn_col = col_letter("Gold FN")
            gold_precision_col = col_letter("Gold Precision")
            gold_recall_col = col_letter("Gold Recall")
            reference_tp_col = col_letter("Reference TP")
            reference_fp_col = col_letter("Reference FP")
            reference_fn_col = col_letter("Reference FN")
            reference_precision_col = col_letter("Reference Precision")
            reference_recall_col = col_letter("Reference Recall")

            set_value("Gold Precision", f"=IFERROR({gold_tp_col}{current_row}/({gold_tp_col}{current_row}+{gold_fp_col}{current_row}),0)")
            set_value("Gold Recall", f"=IFERROR({gold_tp_col}{current_row}/({gold_tp_col}{current_row}+{gold_fn_col}{current_row}),0)")
            set_value("Gold F1", f"=IFERROR(2*{gold_precision_col}{current_row}*{gold_recall_col}{current_row}/({gold_precision_col}{current_row}+{gold_recall_col}{current_row}),0)")
            set_value("Reference Precision", f"=IFERROR({reference_tp_col}{current_row}/({reference_tp_col}{current_row}+{reference_fp_col}{current_row}),0)")
            set_value("Reference Recall", f"=IFERROR({reference_tp_col}{current_row}/({reference_tp_col}{current_row}+{reference_fn_col}{current_row}),0)")
            set_value("Reference F1", f"=IFERROR(2*{reference_precision_col}{current_row}*{reference_recall_col}{current_row}/({reference_precision_col}{current_row}+{reference_recall_col}{current_row}),0)")

            for header in [
                "Dataset Subject Precision",
                "Dataset Subject Recall",
                "Dataset Subject F1",
                "Dataset Temporal Precision",
                "Dataset Temporal Recall",
                "Dataset Temporal F1",
                "Dataset Within 5 mg/dL",
                "Dataset Within 10 mg/dL",
                "Dataset Within 20 mg/dL",
                "Dataset INSIGHT Score",
                "Dataset Robust Score",
                "Scalar Feature Score",
                "Scalar Time In Range Delta",
                "Gold Precision",
                "Gold Recall",
                "Gold F1",
                "Reference Precision",
                "Reference Recall",
                "Reference F1",
            ]:
                ws.cell(current_row, header_col[header]).number_format = "0.00%"

            for header in ["Dataset Glucose MAE", "Dataset Glucose RMSE", "Scalar Mean Glucose Delta", "Scalar GMI Delta"]:
                ws.cell(current_row, header_col[header]).number_format = "0.0000"

            current_row += 1

        sheet_ranges[split] = (split_rows_start, current_row - 1)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, current_row - 1)}"

    widths = {
        "Dataset": 16,
        "Train/Test": 12,
        "# of CGM files in Gold Standard": 16,
        "# of CGM files in Reference Repository": 18,
        "# of files found by INSIGHT": 14,
        "# of files found by INSIGHT in Gold Standard": 16,
        "# of files found by INSIGHT in Reference Repository": 16,
        "# of non-Reference files found by INSIGHT": 16,
        "# of Files that file Loader runs on for Reference Repository files": 18,
        "# of Files that file Loader runs on for non-Reference Repository files": 18,
        "# of files Subject ID correctly extracted for Reference Repository": 16,
        "Reference Repository files extracted correctly": 16,
        "# of files processed": 14,
        "Dataset Subject Precision": 12,
        "Dataset Subject Recall": 12,
        "Dataset Subject F1": 12,
        "Dataset Temporal Precision": 12,
        "Dataset Temporal Recall": 12,
        "Dataset Temporal F1": 12,
        "Dataset Glucose MAE": 12,
        "Dataset Glucose RMSE": 12,
        "Dataset Within 5 mg/dL": 12,
        "Dataset Within 10 mg/dL": 12,
        "Dataset Within 20 mg/dL": 12,
        "Dataset INSIGHT Score": 12,
        "Dataset Robust Score": 12,
        "Scalar Feature Score": 12,
        "Scalar Mean Glucose Delta": 12,
        "Scalar Time In Range Delta": 12,
        "Scalar GMI Delta": 12,
        "Dataset Assessment": 16,
        "Benchmark Status": 16,
        "Benchmark Reason": 36,
        "Reference Coverage Status": 22,
        "Reference Parse Status": 22,
        "Reference Comparison Scope": 28,
        "Notes": 30,
        "Gold Standard Files": 28,
        "Reference Repository Files": 28,
        "INSIGHT Files": 32,
        "INSIGHT Files extracted correctly from Reference Repository": 32,
        "INSIGHT Files processed": 32,
        "Gold TP": 10,
        "Gold FP": 10,
        "Gold FN": 10,
        "Gold Precision": 10,
        "Gold Recall": 10,
        "Gold F1": 10,
        "Reference TP": 10,
        "Reference FP": 10,
        "Reference FN": 10,
        "Reference Precision": 10,
        "Reference Recall": 10,
        "Reference F1": 10,
    }
    for header, width in widths.items():
        ws.column_dimensions[get_column_letter(header_col[header])].width = width

    return sheet_ranges


def build_results_workbook(
    models: list[str],
    runs_root: Path,
    evaluation_root: Path,
    output_path: Path,
    splits: list[str],
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    model_sheet_ranges: dict[str, dict[str, tuple[int, int]]] = {}
    sheet_header_col = {header: index for index, header in enumerate(HEADERS, start=1)}
    summary_header_col = {header: index for index, header in enumerate(SUMMARY_HEADERS, start=1)}

    for model_name in models:
        title = sheet_title_for_model(model_name)
        ws = wb.create_sheet(title)
        model_sheet_ranges[title] = _populate_model_sheet(
            ws=ws,
            model_name=model_name,
            runs_root=runs_root,
            evaluation_root=evaluation_root,
            splits=splits,
        )

    summary = wb.create_sheet("Summary")
    for header, col in summary_header_col.items():
        cell = summary.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = SUMMARY_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    row = 2
    for split in ["training", "testing"]:
        if split not in splits:
            continue
        label = "Train" if split == "training" else "Test"
        for model_name in models:
            title = sheet_title_for_model(model_name)
            split_range = model_sheet_ranges[title].get(split)
            if not split_range:
                continue
            start_row, end_row = split_range

            ws = wb[title]
            active_rows = [
                sheet_row
                for sheet_row in range(start_row, end_row + 1)
                if ws.cell(sheet_row, sheet_header_col["Benchmark Status"]).value == "active"
            ]

            def sum_active(header: str) -> float:
                col = sheet_header_col[header]
                return sum(_safe_number(ws.cell(sheet_row, col).value) for sheet_row in active_rows)

            def avg_active(header: str) -> float:
                col = sheet_header_col[header]
                values = [
                    _safe_number(ws.cell(sheet_row, col).value)
                    for sheet_row in active_rows
                    if ws.cell(sheet_row, col).value is not None
                ]
                return _mean(values)

            active_dataset_count = len(active_rows)
            gold_tp = sum_active("Gold TP")
            gold_fp = sum_active("Gold FP")
            gold_fn = sum_active("Gold FN")
            ref_tp = sum_active("Reference TP")
            ref_fp = sum_active("Reference FP")
            ref_fn = sum_active("Reference FN")
            gold_precision = _safe_rate(gold_tp, gold_tp + gold_fp)
            gold_recall = _safe_rate(gold_tp, gold_tp + gold_fn)
            ref_precision = _safe_rate(ref_tp, ref_tp + ref_fp)
            ref_recall = _safe_rate(ref_tp, ref_tp + ref_fn)

            extracted_correct_count = 0
            for sheet_row in active_rows:
                reference_file_count = _safe_number(
                    ws.cell(sheet_row, sheet_header_col["# of CGM files in Reference Repository"]).value
                )
                extracted_count = _safe_number(
                    ws.cell(sheet_row, sheet_header_col["Reference Repository files extracted correctly"]).value
                )
                if reference_file_count == extracted_count:
                    extracted_correct_count += 1

            summary.cell(row, summary_header_col["Model"], title)
            summary.cell(row, summary_header_col["Train/Test"], label)
            summary.cell(row, summary_header_col["# of Active Datasets"], active_dataset_count)
            summary.cell(row, summary_header_col["# of CGM Files in Gold Standard"], sum_active("# of CGM files in Gold Standard"))
            summary.cell(row, summary_header_col["# of CGM Files in Reference Repository"], sum_active("# of CGM files in Reference Repository"))
            summary.cell(row, summary_header_col["Gold File Precision"], gold_precision)
            summary.cell(row, summary_header_col["Gold File Recall"], gold_recall)
            summary.cell(row, summary_header_col["Gold File F1"], _f1(gold_precision, gold_recall) or 0.0)
            summary.cell(row, summary_header_col["Reference File Precision"], ref_precision)
            summary.cell(row, summary_header_col["Reference File Recall"], ref_recall)
            summary.cell(row, summary_header_col["Reference File F1"], _f1(ref_precision, ref_recall) or 0.0)
            summary.cell(
                row,
                summary_header_col["% of Datasets Extracted Correctly"],
                _safe_rate(extracted_correct_count, active_dataset_count),
            )
            summary.cell(row, summary_header_col["Avg Dataset Subject F1"], avg_active("Dataset Subject F1"))
            summary.cell(row, summary_header_col["Avg Dataset Temporal F1"], avg_active("Dataset Temporal F1"))
            summary.cell(row, summary_header_col["Avg Dataset Within 10 mg/dL"], avg_active("Dataset Within 10 mg/dL"))
            summary.cell(row, summary_header_col["Avg Dataset INSIGHT Score"], avg_active("Dataset INSIGHT Score"))
            summary.cell(row, summary_header_col["Avg Dataset Robust Score"], avg_active("Dataset Robust Score"))
            summary.cell(row, summary_header_col["Avg Scalar Feature Score"], avg_active("Scalar Feature Score"))
            summary.cell(row, summary_header_col["Avg Dataset Glucose MAE"], avg_active("Dataset Glucose MAE"))

            for header in [
                "Gold File Precision",
                "Gold File Recall",
                "Gold File F1",
                "Reference File Precision",
                "Reference File Recall",
                "Reference File F1",
                "% of Datasets Extracted Correctly",
                "Avg Dataset Subject F1",
                "Avg Dataset Temporal F1",
                "Avg Dataset Within 10 mg/dL",
                "Avg Dataset INSIGHT Score",
                "Avg Dataset Robust Score",
                "Avg Scalar Feature Score",
            ]:
                summary.cell(row, summary_header_col[header]).number_format = "0.00%"
            summary.cell(row, summary_header_col["Avg Dataset Glucose MAE"]).number_format = "0.0000"
            row += 1

    for col in range(1, len(SUMMARY_HEADERS) + 1):
        summary.column_dimensions[get_column_letter(col)].width = 24
    summary.freeze_panes = "A2"

    quarantined = wb.create_sheet("Quarantined")
    quarantine_headers = [
        "Model",
        "Train/Test",
        "Dataset",
        "Benchmark Status",
        "Benchmark Reason",
        "Reference Coverage Status",
        "Reference Parse Status",
        "Reference Comparison Scope",
    ]
    for col, header in enumerate(quarantine_headers, start=1):
        cell = quarantined.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = SUMMARY_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    quarantine_row = 2
    for model_name in models:
        slug = model_slug(model_name)
        title = sheet_title_for_model(model_name)
        for split in ["training", "testing"]:
            if split not in splits:
                continue
            for dataset_ref in get_reference_datasets(split):
                benchmark_json = (
                    evaluation_root
                    / slug
                    / "comparisons"
                    / split
                    / dataset_ref.dataset
                    / "benchmark_status.json"
                )
                benchmark_status = _load_benchmark_status(benchmark_json)
                if benchmark_status.get("status") != "quarantined":
                    continue
                quarantined.cell(quarantine_row, 1, title)
                quarantined.cell(quarantine_row, 2, "Train" if split == "training" else "Test")
                quarantined.cell(quarantine_row, 3, dataset_ref.display_name)
                quarantined.cell(quarantine_row, 4, benchmark_status.get("status"))
                quarantined.cell(quarantine_row, 5, benchmark_status.get("reason"))
                quarantined.cell(quarantine_row, 6, benchmark_status.get("reference_coverage_status"))
                quarantined.cell(quarantine_row, 7, benchmark_status.get("reference_parse_status"))
                comparison = _load_comparison_metrics(
                    evaluation_root
                    / slug
                    / "comparisons"
                    / split
                    / dataset_ref.dataset
                    / "comparison.json"
                )
                quarantined.cell(
                    quarantine_row,
                    8,
                    _nested_get(comparison, "reference_scope", "comparison_scope"),
                )
                quarantine_row += 1

    for col in range(1, len(quarantine_headers) + 1):
        quarantined.column_dimensions[get_column_letter(col)].width = 24 if col < 5 else 40
    quarantined.freeze_panes = "A2"

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a Results.xlsx-style workbook from harmony runs.")
    parser.add_argument("--models", nargs="+", default=None)
    parser.add_argument("--runs-root", type=Path, default=Path("harmony/runs"))
    parser.add_argument("--evaluation-root", type=Path, default=Path("harmony/evaluation"))
    parser.add_argument("--output", type=Path, default=Path("Results.gpt54.xlsx"))
    parser.add_argument(
        "--splits",
        nargs="+",
        choices=["training", "testing"],
        default=["training", "testing"],
    )
    args = parser.parse_args()

    models = parse_model_list(args.models) if args.models else list(DEFAULT_MODEL_MATRIX)
    output_path = build_results_workbook(
        models=models,
        runs_root=args.runs_root,
        evaluation_root=args.evaluation_root,
        output_path=args.output,
        splits=args.splits,
    )
    print(output_path)


if __name__ == "__main__":
    main()
